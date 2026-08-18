from fixdata import FIXTURES
from optional_datatypes import require_datatype

XLSXIterable = require_datatype("XLSXIterable")


class TestXLSX:
    def test_id(self):
        datatype_id = XLSXIterable.id()
        assert datatype_id == "xlsx"

    def test_flatonly(self):
        flag = XLSXIterable.is_flatonly()
        assert flag

    def test_openclose(self):
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        iterable.close()

    def test_parsesimple_readone(self):
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        row = iterable.read()
        assert row == FIXTURES[0]
        iterable.close()

    def test_parsesimple_fixedkeys_readone(self):
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx", keys=["id", "name"], start_line=1)
        row = iterable.read()
        assert row == FIXTURES[0]
        iterable.close()

    def test_parsesimple_reset(self):
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        row = iterable.read()
        assert row == FIXTURES[0]
        iterable.reset()
        row_reset = iterable.read()
        assert row_reset == FIXTURES[0]
        iterable.close()

    def test_parsesimple_next(self):
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        row = next(iterable)
        assert row == FIXTURES[0]
        iterable.reset()
        row_reset = next(iterable)
        assert row_reset == FIXTURES[0]
        iterable.close()

    def test_parsesimple_count(self):
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        n = 0
        for _row in iterable:
            n += 1
        assert n == len(FIXTURES)
        iterable.close()

    def test_parsesimple_iterateall(self):
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        n = 0
        for row in iterable:
            assert row == FIXTURES[n]
            n += 1
        iterable.close()

    def test_parsesimple_fixedkeys_iterateall(self):
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx", keys=["id", "name"], start_line=1)
        n = 0
        for row in iterable:
            assert row == FIXTURES[n]
            n += 1
        iterable.close()

    def test_has_tables(self):
        """Test has_tables static method"""
        assert XLSXIterable.has_tables() is True

    def test_list_tables_instance_method(self):
        """Test list_tables on an already-opened instance"""
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        sheets = iterable.list_tables()
        assert isinstance(sheets, list)
        assert len(sheets) > 0
        # Sheet name may be localized (e.g., "Sheet1", "Лист1"), just check it's a string
        assert isinstance(sheets[0], str)
        assert len(sheets[0]) > 0
        iterable.close()

    def test_list_tables_with_filename(self):
        """Test list_tables with filename parameter (class-like usage)"""
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        sheets = iterable.list_tables("fixtures/2cols6rows.xlsx")
        assert isinstance(sheets, list)
        assert len(sheets) > 0
        iterable.close()

    def test_list_tables_multiple_sheets(self):
        """Test list_tables with file containing multiple sheets"""
        iterable = XLSXIterable("fixtures/multi_sheet.xlsx")
        sheets = iterable.list_tables()
        assert isinstance(sheets, list)
        assert len(sheets) >= 3  # Should have at least Sheet, Sheet2, Data
        assert "Sheet" in sheets or "Sheet1" in sheets
        assert "Sheet2" in sheets
        assert "Data" in sheets
        iterable.close()

    def test_list_tables_reuses_workbook(self):
        """Test that list_tables reuses open workbook"""
        iterable = XLSXIterable("fixtures/2cols6rows.xlsx")
        # Read a row to ensure workbook is open
        _ = iterable.read()
        # Now list tables - should reuse workbook
        sheets1 = iterable.list_tables()
        sheets2 = iterable.list_tables()
        assert sheets1 == sheets2
        iterable.close()

    def test_list_tables_empty_file(self):
        """Test list_tables on empty Excel file (edge case)"""
        import os
        import tempfile

        from openpyxl import Workbook

        # Create empty workbook
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_file = f.name

        try:
            wb = Workbook()
            wb.save(temp_file)
            wb.close()

            iterable = XLSXIterable(temp_file)
            sheets = iterable.list_tables()
            # Empty workbook should still have at least one sheet
            assert isinstance(sheets, list)
            iterable.close()
        finally:
            if os.path.exists(temp_file):
                os.unlink(temp_file)

    def test_open_table_uses_page_index_not_ignored_table_option(self, tmp_path):
        """Excel must open the named sheet; table= must not silently keep page 0."""

        from openpyxl import Workbook

        from iterable.ai.fileinfo import open_table

        path = tmp_path / "sheets.xlsx"
        book = Workbook()
        book.active.title = "Disclaimer"
        book.active["A1"] = "notice"
        data = book.create_sheet("Data")
        data["A1"] = "id"
        data["B1"] = "name"
        data["A2"] = 1
        data["B2"] = "alpha"
        book.save(path)

        opened = open_table(str(path), "Data")
        try:
            assert opened.page == 1
            assert opened.keys == ["id", "name"]
            assert opened.read()["name"] == "alpha"
        finally:
            opened.close()

    def test_reset_dimensions_reads_past_wrong_a1_dimension(self, tmp_path):
        """Workbooks with dimension ref=A1 must still stream full rows."""

        import re
        import zipfile

        from openpyxl import Workbook

        raw = tmp_path / "raw.xlsx"
        path = tmp_path / "bad_dims.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Data"
        sheet.append(["Entity", "Week", "Flights"])
        sheet.append(["Albania", 40, 214])
        sheet.append(["Belgium", 41, 300])
        book.save(raw)

        with zipfile.ZipFile(raw, "r") as zin, zipfile.ZipFile(path, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/worksheets/sheet") and b"Entity" in data:
                    text = data.decode()
                    text, count = re.subn(
                        r'<dimension ref="[^"]*"/>',
                        '<dimension ref="A1"/>',
                        text,
                        count=1,
                    )
                    assert count == 1
                    data = text.encode()
                zout.writestr(item, data)

        iterable = XLSXIterable(str(path))
        try:
            assert iterable.keys == ["Entity", "Week", "Flights"]
            assert iterable.read()["Entity"] == "Albania"
            assert iterable.read()["Entity"] == "Belgium"
        finally:
            iterable.close()

    def test_skips_leading_blank_rows_before_header(self, tmp_path):
        from openpyxl import Workbook

        path = tmp_path / "leading_blank.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Metadata"
        sheet["A2"] = "Column name"
        sheet["B2"] = "Description"
        sheet["A3"] = "Entity"
        sheet["B3"] = "Name of the State"
        book.save(path)

        iterable = XLSXIterable(str(path))
        try:
            assert iterable.keys == ["Column name", "Description"]
            assert iterable.read()["Column name"] == "Entity"
        finally:
            iterable.close()
